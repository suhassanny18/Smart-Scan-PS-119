"""
Smart Attendance System — Flask Backend (v3)
Three-tier access control:
  admin       → Full system control
  dept_head   → Manages dept: faculty, timetable, sessions, leave/sub
  faculty     → Calendar view of sessions, mark attendance

Persistent store: system_db.json
Attendance log:   attendance_log.csv
"""

from flask import Flask, Response, jsonify, request, session
import cv2, numpy as np, csv, json, datetime, threading, time, os, hashlib, io, base64
import google.generativeai as genai

try:
    from insightface.app import FaceAnalysis
    INSIGHTFACE_AVAILABLE = True
except ImportError:
    INSIGHTFACE_AVAILABLE = False
    print("⚠️  insightface not installed – face recognition disabled (demo mode)")

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ── Constants ─────────────────────────────────────────────────────────────────
KNOWN_FACES_DIR = "known_faces"
CSV_FILE        = "attendance_log.csv"
DB_FILE         = "system_db.json"
SECRET_KEY      = "your-random-flask-secret-2025"
GOOGLE_API_KEY  = "AIzaSyBpElFLwdXIj0nROmuWGxtCgKPLk0igtIM"

DEPARTMENTS = ["CSE", "ECE", "EEE", "MECH", "CIVIL"]
ALL_SECTIONS = {
    "CSE":   [f"CSE-{i}" for i in range(1, 10)],
    "ECE":   [f"ECE-{i}" for i in range(1, 7)],
    "EEE":   [f"EEE-{i}" for i in range(1, 5)],
    "MECH":  [f"MECH-{i}" for i in range(1, 5)],
    "CIVIL": [f"CIVIL-{i}" for i in range(1, 4)],
}

# Timetable slots 9:30–4:30, each 50 mins
TIME_SLOTS = [
    {"id": "slot1",  "label": "Period 1",  "start": "09:30", "end": "10:20"},
    {"id": "slot2",  "label": "Period 2",  "start": "10:20", "end": "11:10"},
    {"id": "slot3",  "label": "Period 3",  "start": "11:10", "end": "12:00"},
    {"id": "slot4",  "label": "Lunch",     "start": "12:00", "end": "12:40"},
    {"id": "slot5",  "label": "Period 4",  "start": "12:40", "end": "13:30"},
    {"id": "slot6",  "label": "Period 5",  "start": "13:30", "end": "14:20"},
    {"id": "slot7",  "label": "Period 6",  "start": "14:20", "end": "15:10"},
    {"id": "slot8",  "label": "Period 7",  "start": "15:10", "end": "16:00"},
    {"id": "slot9",  "label": "Period 8",  "start": "16:00", "end": "16:30"},
]
BREAK_SLOTS = {"slot4"}

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

CAMERA_SOURCES = [
    {
        "id": "default",
        "label": "Device Camera",
        "source": 0          # built-in / USB webcam (index 0)
    },
]

# ── Flask app ─────────────────────────────────────────────────────────────────
app = Flask(__name__)
genai.configure(api_key=GOOGLE_API_KEY)
app.secret_key = SECRET_KEY

# ── In-memory state ───────────────────────────────────────────────────────────
frame_lock    = threading.Lock()
camera_frames = {}
_state_lock   = threading.Lock()
attendance_today = {}

face_app         = None
known_embeddings = {}

# ── Active automated sessions ─────────────────────────────────────────────────
# key: session_id → {thread, stop_event, captures, detected_counts, start_time,
#                    end_time, section, slot_id, dept, faculty, cam_id, date}
active_sessions: dict = {}
_session_lock = threading.Lock()

# ═══════════════════════════════════════════════════════════════════════════════
#  DATABASE
# ═══════════════════════════════════════════════════════════════════════════════
def _default_db():
    return {
        "users": {
            "admin": {
                "password_hash": _hash("admin123"),
                "role": "admin",
                "department": None,
                "name": "System Administrator",
                "sections": []
            }
        },
        "timetable": {},
        "students":  {s: [] for dept in ALL_SECTIONS for s in ALL_SECTIONS[dept]},
        "camera_map": {s: "default" for dept in ALL_SECTIONS for s in ALL_SECTIONS[dept]},
        "sessions": [],
        "attendance_records": {},
    }

def _hash(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

def load_db() -> dict:
    default = _default_db()
    if os.path.exists(DB_FILE):
        try:
            with open(DB_FILE) as f:
                saved = json.load(f)
            for k, v in default.items():
                if k not in saved:
                    saved[k] = v
                elif isinstance(v, dict):
                    for k2, v2 in v.items():
                        saved[k].setdefault(k2, v2)
            return saved
        except Exception as e:
            print(f"⚠️  DB load error: {e} — using default")
    return default

def save_db():
    with open(DB_FILE, "w") as f:
        json.dump(db, f, indent=2)

db = load_db()

# ── CSV ───────────────────────────────────────────────────────────────────────
def setup_csv():
    if not os.path.exists(CSV_FILE):
        with open(CSV_FILE, "w", newline="") as f:
            csv.writer(f).writerow(
                ["Roll", "Name", "Section", "Department", "Slot", "Subject",
                 "Faculty", "Date", "Time", "Status"])

def write_attendance_csv(roll, name, section, dept, slot_id, subject, faculty, status="Present"):
    """Append one row to the CSV log. Duplicates are possible if the same session
    is saved multiple times; use the JSON attendance_records as the authoritative
    source of truth and treat the CSV as an audit trail only."""
    now = datetime.datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")
    with open(CSV_FILE, "a", newline="") as f:
        csv.writer(f).writerow(
            [roll, name, section, dept, slot_id, subject, faculty, date_str, time_str, status])

def rewrite_attendance_csv_for_key(section, dept, slot_id, date_str, records, faculty):
    """Rewrite ALL rows for a specific section+slot+date in the CSV to avoid duplicates.
    Called by faculty_save_attendance and dept_update_attendance on every manual save."""
    CANONICAL_FIELDS = ["Roll","Name","Section","Department","Slot","Subject","Faculty","Date","Time","Status"]
    if not os.path.exists(CSV_FILE):
        setup_csv()
    # Read existing rows that are NOT for this key
    kept_rows = []
    try:
        with open(CSV_FILE, newline="") as f:
            reader = csv.DictReader(f)
            old_fields = reader.fieldnames or []
            # Only keep old rows if the CSV has the canonical headers
            if set(CANONICAL_FIELDS).issubset(set(old_fields)):
                for row in reader:
                    if not (row.get("Section") == section and row.get("Slot") == slot_id and row.get("Date") == date_str):
                        kept_rows.append({k: row.get(k,"") for k in CANONICAL_FIELDS})
    except Exception:
        pass
    # Rewrite file with canonical headers + kept rows + fresh records for this key
    time_str = datetime.datetime.now().strftime("%H:%M:%S")
    with open(CSV_FILE, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CANONICAL_FIELDS)
        w.writeheader()
        w.writerows(kept_rows)
        for r in records:
            w.writerow({
                "Roll": r["roll"], "Name": r.get("name",""), "Section": section,
                "Department": dept, "Slot": slot_id, "Subject": "",
                "Faculty": faculty, "Date": date_str, "Time": time_str,
                "Status": r.get("status","Absent"),
            })

# ═══════════════════════════════════════════════════════════════════════════════
#  SESSION / TIMETABLE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def current_slot_id():
    now = datetime.datetime.now().strftime("%H:%M")
    for slot in TIME_SLOTS:
        if slot["start"] <= now < slot["end"]:
            return slot["id"]
    return None

def current_day() -> str:
    return datetime.datetime.now().strftime("%A")

def generate_sessions_for_week(faculty_username: str, dept: str, week_offset: int = 0):
    tt = db["timetable"].get(faculty_username, {})
    if not tt:
        return
    today = datetime.date.today()
    monday = today - datetime.timedelta(days=today.weekday()) + datetime.timedelta(weeks=week_offset)
    day_offsets = {d: i for i, d in enumerate(DAYS)}

    # Build index of existing sessions by composite key for fast lookup
    existing_by_key = {}
    for s in db["sessions"]:
        k = f"{s['faculty_username']}::{s['date']}::{s['slot_id']}::{s['section']}"
        existing_by_key[k] = s

    new_sessions = []
    changed = False
    for day_name, slots in tt.items():
        offset = day_offsets.get(day_name)
        if offset is None:
            continue
        date_obj = monday + datetime.timedelta(days=offset)
        date_str = date_obj.isoformat()
        for slot_id, entry in slots.items():
            if slot_id in BREAK_SLOTS:
                continue
            section = entry.get("section", "")
            subject  = entry.get("subject", "")
            if not section:
                continue
            key = f"{faculty_username}::{date_str}::{slot_id}::{section}"
            slot_info = next((s for s in TIME_SLOTS if s["id"] == slot_id), {})
            if key in existing_by_key:
                # Update mutable fields so dept_head edits always reflect
                sess = existing_by_key[key]
                updated = False
                if sess.get("subject") != subject:
                    sess["subject"] = subject; updated = True
                if sess.get("dept") != dept:
                    sess["dept"] = dept; updated = True
                if sess.get("slot_start") != slot_info.get("start", ""):
                    sess["slot_start"] = slot_info.get("start", ""); updated = True
                if sess.get("slot_end") != slot_info.get("end", ""):
                    sess["slot_end"] = slot_info.get("end", ""); updated = True
                # Always keep att_key in sync with section+slot+date
                expected_att_key = f"{section}::{slot_id}::{date_str}"
                if sess.get("att_key") != expected_att_key:
                    sess["att_key"] = expected_att_key; updated = True
                if updated:
                    changed = True
            else:
                new_sessions.append({
                    "id": f"{faculty_username}_{date_str}_{slot_id}_{section}",
                    "date": date_str,
                    "section": section,
                    "slot_id": slot_id,
                    "slot_label": slot_info.get("label", slot_id),
                    "slot_start": slot_info.get("start", ""),
                    "slot_end": slot_info.get("end", ""),
                    "subject": subject,
                    "faculty_username": faculty_username,
                    "dept": dept,
                    "status": "scheduled",
                    "substitute": None,
                    "att_key": f"{section}::{slot_id}::{date_str}",
                })

    if new_sessions:
        db["sessions"].extend(new_sessions)
        changed = True
    if changed:
        save_db()

def get_faculty_sessions(faculty_username: str, week_offset: int = 0):
    today = datetime.date.today()
    monday = today - datetime.timedelta(days=today.weekday()) + datetime.timedelta(weeks=week_offset)
    sunday = monday + datetime.timedelta(days=6)
    result = []
    for s in db["sessions"]:
        if s.get("faculty_username") != faculty_username and s.get("substitute") != faculty_username:
            continue
        try:
            d = datetime.date.fromisoformat(s["date"])
        except Exception:
            continue
        if monday <= d <= sunday:
            result.append(s)
    return result

# ═══════════════════════════════════════════════════════════════════════════════
#  PHOTO HELPER
# ═══════════════════════════════════════════════════════════════════════════════
def get_student_photo_b64(roll: str):
    for ext in (".jpg", ".jpeg", ".png"):
        path = os.path.join(KNOWN_FACES_DIR, roll + ext)
        if os.path.exists(path):
            try:
                with open(path, "rb") as f:
                    data = base64.b64encode(f.read()).decode()
                mime = "image/jpeg" if ext in (".jpg", ".jpeg") else "image/png"
                return f"data:{mime};base64,{data}"
            except Exception:
                pass
    return None

# ═══════════════════════════════════════════════════════════════════════════════
#  CAMERA
# ═══════════════════════════════════════════════════════════════════════════════
def camera_loop(cam_id: str, source):
    cap = cv2.VideoCapture(source)
    cap.set(cv2.CAP_PROP_BUFFERSIZE, 2)
    if not cap.isOpened():
        print(f"❌ Cannot open camera {cam_id}")
        return
    print(f"📷 Camera ready: {cam_id}")
    while True:
        ret, frame = cap.read()
        if not ret:
            time.sleep(0.1)
            continue
        if INSIGHTFACE_AVAILABLE and face_app:
            try:
                for face in face_app.get(frame):
                    x1,y1,x2,y2 = face.bbox.astype(int)
                    cv2.rectangle(frame,(x1,y1),(x2,y2),(0,200,80),2)
            except Exception:
                pass
        with frame_lock:
            camera_frames[cam_id] = frame.copy()
        time.sleep(0.03)

def gen_frames(cam_id: str):
    while True:
        with frame_lock:
            frame = camera_frames.get(cam_id)
        if frame is None:
            time.sleep(0.05)
            continue
        ok, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 80])
        if not ok:
            continue
        yield b"--frame\r\nContent-Type: image/jpeg\r\n\r\n" + buf.tobytes() + b"\r\n"
        time.sleep(0.04)

# ═══════════════════════════════════════════════════════════════════════════════
#  FACE RECOGNITION
# ═══════════════════════════════════════════════════════════════════════════════
def init_face_model():
    global face_app
    if not INSIGHTFACE_AVAILABLE:
        return
    print("🤖 Loading InsightFace …")
    face_app = FaceAnalysis(name="buffalo_l",
                             providers=["CoreMLExecutionProvider","CPUExecutionProvider"])
    face_app.prepare(ctx_id=0, det_size=(640,640))
    print("✓ InsightFace ready")

def load_known_faces():
    if not (INSIGHTFACE_AVAILABLE and face_app):
        return
    os.makedirs(KNOWN_FACES_DIR, exist_ok=True)
    for fname in os.listdir(KNOWN_FACES_DIR):
        if not fname.lower().endswith((".png",".jpg",".jpeg")):
            continue
        img = cv2.imread(os.path.join(KNOWN_FACES_DIR, fname))
        if img is None:
            continue
        try:
            faces = face_app.get(img)
            if faces:
                name = os.path.splitext(fname)[0]
                known_embeddings[name] = faces[0].embedding
                print(f"  ✓ {name}")
        except Exception as e:
            print(f"  ❌ {fname}: {e}")

def cosine_dist(e1, e2):
    return 1 - np.dot(e1,e2)/(np.linalg.norm(e1)*np.linalg.norm(e2)+1e-9)

def match_face(emb):
    best, dist = None, float("inf")
    for name, ref in known_embeddings.items():
        d = cosine_dist(emb, ref)
        if d < dist:
            dist, best = d, name
    return best, dist

def liveness(f1, f2):
    if not (INSIGHTFACE_AVAILABLE and face_app):
        return True, 0.5
    try:
        fa1,fa2 = face_app.get(f1), face_app.get(f2)
        if not fa1 or not fa2:
            return False, 0.0
        lm1,lm2 = fa1[0].landmark_2d_106, fa2[0].landmark_2d_106
        if lm1 is None or lm2 is None:
            return True, 0.5
        pts=[28,30,35,40,52,64]
        ds=[np.linalg.norm(lm2[i]-lm1[i]) for i in pts if i<len(lm1)]
        avg=float(np.mean(ds)) if ds else 0.0
        return avg>1.5, avg
    except Exception:
        return True, 0.5

def capture_frame_detections(cam_id: str) -> set:
    """Capture one frame and return a set of matched student names."""
    TOLE = 0.6
    if not (INSIGHTFACE_AVAILABLE and face_app):
        return set()
    with frame_lock:
        frame = camera_frames.get(cam_id)
    if frame is None:
        return set()
    frame = frame.copy()
    try:
        faces = face_app.get(frame)
        detected = set()
        for face in faces:
            name, dist = match_face(face.embedding)
            if name and dist < TOLE:
                detected.add(name)
        return detected
    except Exception:
        return set()


def session_runner(session_id: str):
    """
    Background thread that auto-captures attendance at random ~10-min intervals.
    Finalizes 1 minute before class ends.
    A student is marked Present if detected in ≥2 captures.
    """
    with _session_lock:
        info = active_sessions.get(session_id)
    if not info:
        return

    stop_evt      = info["stop_event"]
    end_time      = info["end_time"]          # datetime
    section       = info["section"]
    slot_id       = info["slot_id"]
    dept          = info["dept"]
    faculty       = info["faculty"]
    cam_id        = info["cam_id"]
    date_str      = info["date"]
    stu_name_to_roll = {s["name"]: s["roll"] for s in db["students"].get(section, [])}

    finalize_at = end_time - datetime.timedelta(minutes=1)
    INTERVAL_BASE = 10 * 60   # 10 minutes in seconds
    INTERVAL_JITTER = 90      # ±90 seconds

    def do_capture():
        detected = capture_frame_detections(cam_id)
        with _session_lock:
            sess = active_sessions.get(session_id)
            if sess is None:
                return
            sess["captures"] += 1
            for name in detected:
                sess["detected_counts"][name] = sess["detected_counts"].get(name, 0) + 1

    def finalize():
        with _session_lock:
            sess = active_sessions.get(session_id)
            if sess is None:
                return
            sess["status"] = "finalizing"
            detected_counts = dict(sess["detected_counts"])

        # Mark present if detected in ≥2 captures
        present_names = {name for name, cnt in detected_counts.items() if cnt >= 2}
        all_students  = db["students"].get(section, [])
        key = f"{section}::{slot_id}::{date_str}"
        now_str = datetime.datetime.now().strftime("%H:%M:%S")

        with _state_lock:
            # Build a fresh authoritative list — upsert every student
            roll_to_rec = {}
            for r in db["attendance_records"].get(key, []):
                roll_to_rec[r["roll"]] = r

            for stu in all_students:
                roll   = stu["roll"]
                name   = stu["name"]
                status = "Present" if name in present_names else "Absent"
                roll_to_rec[roll] = {
                    "roll": roll, "name": name, "status": status,
                    "time": now_str, "marked_by": "auto_session"
                }

            db["attendance_records"][key] = list(roll_to_rec.values())
        # Rewrite CSV for this key to avoid duplicates from multiple finalize attempts
        rewrite_attendance_csv_for_key(section, dept, slot_id, date_str,
                                       db["attendance_records"][key], faculty)
        save_db()

        with _session_lock:
            sess = active_sessions.get(session_id)
            if sess:
                sess["status"] = "completed"

    # ── main loop ──────────────────────────────────────────────────────────────
    # Do an initial capture shortly after start (30–90 s)
    first_wait = 30 + (hash(session_id) % 60)
    stop_evt.wait(timeout=first_wait)
    if stop_evt.is_set():
        finalize(); return
    do_capture()

    import random
    while True:
        now = datetime.datetime.now()
        if now >= finalize_at:
            finalize()
            return

        # Sleep until next capture or finalize time
        next_capture = now + datetime.timedelta(
            seconds=INTERVAL_BASE + random.randint(-INTERVAL_JITTER, INTERVAL_JITTER))
        sleep_until  = min(next_capture, finalize_at)
        secs_to_sleep = (sleep_until - datetime.datetime.now()).total_seconds()
        if secs_to_sleep > 0:
            stop_evt.wait(timeout=secs_to_sleep)
        if stop_evt.is_set():
            finalize(); return
        now = datetime.datetime.now()
        if now >= finalize_at:
            finalize(); return
        do_capture()

# ═══════════════════════════════════════════════════════════════════════════════
#  AUTH HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def me():
    """Return user dict or None (never empty dict, so `if not u` works correctly)."""
    uname = session.get("username")
    if not uname:
        return None
    return db["users"].get(uname) or None

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — AUTH
# ═══════════════════════════════════════════════════════════════════════════════
@app.route("/login", methods=["POST"])
def api_login():
    d = request.get_json(silent=True) or {}
    uname = d.get("username","").strip()
    pwd   = d.get("password","")
    user  = db["users"].get(uname)
    if not user or user["password_hash"] != _hash(pwd):
        return jsonify({"success": False, "message": "Invalid credentials"}), 401
    session.update(logged_in=True, username=uname,
                   role=user["role"], department=user.get("department"),
                   name=user.get("name",""))
    return jsonify({
        "success":    True,
        "role":       user["role"],
        "department": user.get("department"),
        "name":       user.get("name",""),
        "sections":   user.get("sections", []),
        "username":   uname,
    })

@app.route("/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"success": True})

@app.route("/check_session")
def api_check_session():
    u = me()
    if not u:
        return jsonify({"logged_in": False})
    return jsonify({
        "logged_in":  True,
        "role":       u.get("role"),
        "department": u.get("department"),
        "name":       u.get("name",""),
        "sections":   u.get("sections",[]),
        "username":   session.get("username"),
    })

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — CAMERA
# ═══════════════════════════════════════════════════════════════════════════════
@app.route("/cameras")
def api_cameras():
    if not me():
        return jsonify({"cameras": []}), 401
    return jsonify({"cameras": CAMERA_SOURCES})

@app.route("/video_feed")
def video_feed():
    if not me():
        return Response("Unauthorised", status=401)
    cam_id = request.args.get("cam", "default")
    return Response(gen_frames(cam_id),
                    mimetype="multipart/x-mixed-replace; boundary=frame")

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — FACULTY
# ═══════════════════════════════════════════════════════════════════════════════
@app.route("/faculty/calendar")
def faculty_calendar():
    u = me()
    if not u:
        return jsonify({"sessions": []}), 401
    week_offset = int(request.args.get("week", 0))
    uname = session.get("username")
    dept = u.get("department") or ""
    if not dept:
        for s in db["sessions"]:
            if s.get("faculty_username") == uname and s.get("dept"):
                dept = s["dept"]
                break
    # Generate sessions for the requested week
    generate_sessions_for_week(uname, dept, week_offset)
    sessions = get_faculty_sessions(uname, week_offset)
    for s in sessions:
        key = s.get("att_key", "")
        recs = db["attendance_records"].get(key, [])
        all_stu = db["students"].get(s["section"], [])
        total = len(all_stu)
        present = sum(1 for r in recs if r["status"] == "Present")
        s["att_count"] = present
        s["total_students"] = total
        # att_marked: true only when every student has a record
        s["att_marked"] = len(recs) >= total > 0
        orig_user = db["users"].get(s.get("faculty_username", ""), {})
        s["faculty_name"] = orig_user.get("name", s.get("faculty_username", ""))
        if s.get("substitute"):
            sub_user = db["users"].get(s["substitute"], {})
            s["substitute_name"] = sub_user.get("name", s["substitute"])
        else:
            s["substitute_name"] = None
    return jsonify({"sessions": sessions, "time_slots": TIME_SLOTS})

@app.route("/faculty/session_detail")
def faculty_session_detail():
    u = me()
    if not u:
        return jsonify({}), 401
    session_id = request.args.get("id","")
    sess = next((s for s in db["sessions"] if s["id"] == session_id), None)
    if not sess:
        return jsonify({"error": "Session not found"}), 404
    uname = session.get("username")
    if sess["faculty_username"] != uname and sess.get("substitute") != uname:
        if u.get("role") not in ("admin","dept_head"):
            return jsonify({"error": "Forbidden"}), 403
    key = sess.get("att_key","")
    recs = db["attendance_records"].get(key, [])
    all_students = db["students"].get(sess["section"], [])
    present_rolls = {r["roll"] for r in recs if r["status"] == "Present"}
    student_list = []
    for stu in all_students:
        student_list.append({
            "roll": stu["roll"],
            "name": stu["name"],
            "status": "Present" if stu["roll"] in present_rolls else "Absent",
            "photo": get_student_photo_b64(stu["roll"]),
        })
    return jsonify({
        "session": sess,
        "students": student_list,
        "present_count": len(present_rolls),
        "total": len(all_students),
    })

@app.route("/faculty/start_session", methods=["POST"])
def faculty_start_session():
    u = me()
    if not u:
        return jsonify({"success": False, "message": "Not logged in"}), 401
    d          = request.get_json(silent=True) or {}
    session_id = d.get("session_id", "")
    cam_id     = d.get("cam_id", "default")

    sess = next((s for s in db["sessions"] if s["id"] == session_id), None)
    if not sess:
        return jsonify({"success": False, "message": "Session not found"}), 404

    uname = session.get("username")
    if sess["faculty_username"] != uname and sess.get("substitute") != uname:
        if u.get("role") not in ("admin", "dept_head"):
            return jsonify({"success": False, "message": "Forbidden"}), 403

    # already running?
    with _session_lock:
        if session_id in active_sessions and active_sessions[session_id]["status"] in ("running", "finalizing"):
            info = active_sessions[session_id]
            return jsonify({
                "success": True,
                "already_running": True,
                "captures": info["captures"],
                "status": info["status"],
                "end_time": info["end_time"].strftime("%H:%M"),
                "finalize_at": (info["end_time"] - datetime.timedelta(minutes=1)).strftime("%H:%M"),
            })

    # parse slot end time to build end datetime
    slot_end_str = sess.get("slot_end", "")
    sess_date    = sess.get("date", datetime.date.today().isoformat())
    try:
        h, m = map(int, slot_end_str.split(":"))
        date_obj  = datetime.date.fromisoformat(sess_date)
        end_dt    = datetime.datetime.combine(date_obj, datetime.time(h, m))
    except Exception:
        return jsonify({"success": False, "message": "Could not parse slot end time"}), 400

    # If the class period is already over, still allow manual attendance marking
    # but don't start the auto-capture thread
    now = datetime.datetime.now()
    if now >= end_dt:
        # Allow session to be "started" for manual attendance even if time passed
        # Mark it completed immediately so manual save is still possible
        return jsonify({
            "success": False,
            "message": "Class period has ended. Use manual attendance marking below.",
            "period_ended": True,
        }), 200

    stop_evt = threading.Event()
    info = {
        "session_id": session_id,
        "section":    sess["section"],
        "slot_id":    sess["slot_id"],
        "dept":       sess.get("dept", u.get("department", "")),
        "faculty":    u.get("name", ""),
        "cam_id":     cam_id,
        "date":       sess_date,
        "end_time":   end_dt,
        "stop_event": stop_evt,
        "captures":   0,
        "detected_counts": {},
        "status":     "running",
        "started_at": now.isoformat(),
    }
    with _session_lock:
        active_sessions[session_id] = info

    t = threading.Thread(target=session_runner, args=(session_id,), daemon=True)
    info["thread"] = t
    t.start()

    return jsonify({
        "success": True,
        "message": "Session started — attendance will be captured automatically",
        "end_time": end_dt.strftime("%H:%M"),
        "finalize_at": (end_dt - datetime.timedelta(minutes=1)).strftime("%H:%M"),
    })


@app.route("/faculty/session_status")
def faculty_session_status():
    u = me()
    if not u:
        return jsonify({}), 401
    session_id = request.args.get("id", "")
    with _session_lock:
        info = active_sessions.get(session_id)
    if not info:
        return jsonify({"active": False})

    section  = info["section"]
    slot_id  = info["slot_id"]
    date_str = info["date"]
    key      = f"{section}::{slot_id}::{date_str}"
    recs     = db["attendance_records"].get(key, [])
    total    = len(db["students"].get(section, []))
    present  = sum(1 for r in recs if r["status"] == "Present")

    # compute minutes remaining
    now = datetime.datetime.now()
    end_time = info["end_time"]
    mins_left = max(0, int((end_time - now).total_seconds() // 60))

    return jsonify({
        "active":    info["status"] in ("running", "finalizing"),
        "status":    info["status"],
        "captures":  info["captures"],
        "mins_left": mins_left,
        "end_time":  end_time.strftime("%H:%M"),
        "present":   present,
        "total":     total,
        "detected_preview": [
            {"name": name, "count": cnt}
            for name, cnt in sorted(info["detected_counts"].items(), key=lambda x: -x[1])
        ][:10],
    })

@app.route("/faculty/save_attendance", methods=["POST"])
def faculty_save_attendance():
    u = me()
    if not u:
        return jsonify({"success": False}), 401
    d        = request.get_json(silent=True) or {}
    section  = d.get("section", "")
    slot_id  = d.get("slot_id", "")
    date_str = d.get("date", datetime.date.today().isoformat())
    records  = d.get("records", [])
    dept     = u.get("department", "") or ""

    if not section or not slot_id:
        return jsonify({"success": False, "error": "section and slot_id required"}), 400

    # Normalise records: ensure each has roll, name, status
    clean_records = []
    for r in records:
        if not r.get("roll"):
            continue
        clean_records.append({
            "roll":      r["roll"],
            "name":      r.get("name", r["roll"]),
            "status":    r.get("status", "Absent"),
            "time":      datetime.datetime.now().strftime("%H:%M:%S"),
            "marked_by": session.get("username", "faculty"),
        })

    key = f"{section}::{slot_id}::{date_str}"
    with _state_lock:
        db["attendance_records"][key] = clean_records
    save_db()
    # Use rewrite to avoid duplicate CSV rows on re-saves
    rewrite_attendance_csv_for_key(section, dept, slot_id, date_str, clean_records, u.get("name",""))
    return jsonify({"success": True, "saved": len(clean_records)})

@app.route("/faculty/attendance_list")
def faculty_attendance_list():
    if not me():
        return jsonify({"records": []}), 401
    section  = request.args.get("section","")
    slot_id  = request.args.get("slot","")
    date_str = request.args.get("date", datetime.date.today().isoformat())
    key = f"{section}::{slot_id}::{date_str}"
    recs = db["attendance_records"].get(key, [])
    for r in recs:
        if "photo" not in r:
            r["photo"] = get_student_photo_b64(r["roll"])
    return jsonify({"records": recs, "count": len(recs)})

@app.route("/faculty/slot_info")
def faculty_slot_info():
    if not me():
        return jsonify({}), 401
    sid = current_slot_id()
    day = current_day()
    return jsonify({
        "slot_id":    sid,
        "day":        day,
        "slot_label": next((s["label"] for s in TIME_SLOTS if s["id"]==sid), "—"),
        "slot_time":  next((f"{s['start']}–{s['end']}" for s in TIME_SLOTS if s["id"]==sid), ""),
        "time_slots": TIME_SLOTS,
    })

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — DEPT HEAD
# ═══════════════════════════════════════════════════════════════════════════════
@app.route("/dept/faculty_list")
def dept_faculty_list():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    dept = u["department"] if u["role"]=="dept_head" else request.args.get("dept","CSE")
    users = [
        {"username": un, "name": ud["name"], "role": ud["role"],
         "sections": ud.get("sections",[]), "department": ud.get("department")}
        for un, ud in db["users"].items()
        if ud.get("department") == dept and ud["role"] in ("faculty","dept_head")
    ]
    return jsonify({"faculty": users, "sections": ALL_SECTIONS.get(dept,[])})

@app.route("/dept/create_faculty", methods=["POST"])
def dept_create_faculty():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d = request.get_json(silent=True) or {}
    uname    = d.get("username","").strip()
    password = d.get("password","")
    name     = d.get("name","")
    role     = d.get("role","faculty")
    dept     = d.get("department","")
    sections = d.get("sections",[])

    if u["role"]=="dept_head":
        dept = u["department"]
        if role == "admin":
            return jsonify({"error":"Cannot create admin"}),403
    if not uname or not password:
        return jsonify({"error":"Username and password required"}),400
    if uname in db["users"]:
        return jsonify({"error":"Username already exists"}),409

    db["users"][uname] = {
        "password_hash": _hash(password),
        "role": role,
        "department": dept,
        "name": name or uname,
        "sections": sections,
    }
    save_db()
    return jsonify({"success": True})

@app.route("/dept/update_faculty", methods=["POST"])
def dept_update_faculty():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d = request.get_json(silent=True) or {}
    uname = d.get("username","")
    if uname not in db["users"]:
        return jsonify({"error":"User not found"}),404
    target = db["users"][uname]
    if u["role"]=="dept_head" and target.get("department")!=u["department"]:
        return jsonify({"error":"Cross-department forbidden"}),403
    if "name"     in d: target["name"]     = d["name"]
    if "sections" in d: target["sections"] = d["sections"]
    if "role"     in d and u["role"]=="admin": target["role"] = d["role"]
    if "password" in d and d["password"]: target["password_hash"] = _hash(d["password"])
    save_db()
    return jsonify({"success": True})

@app.route("/dept/delete_faculty", methods=["POST"])
def dept_delete_faculty():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d = request.get_json(silent=True) or {}
    uname = d.get("username","")
    if uname == "admin":
        return jsonify({"error":"Cannot delete root admin"}),403
    if uname not in db["users"]:
        return jsonify({"error":"Not found"}),404
    del db["users"][uname]
    save_db()
    return jsonify({"success": True})

# ── Timetable ──────────────────────────────────────────────────────────────────
@app.route("/dept/timetable")
def dept_timetable():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    dept = u["department"] if u["role"]=="dept_head" else request.args.get("dept","CSE")
    faculty_username = request.args.get("faculty","")
    day = request.args.get("day","Monday")
    tt = db["timetable"].get(faculty_username, {}).get(day, {}) if faculty_username else {}
    faculty_list = [
        {"username": un, "name": ud["name"]}
        for un, ud in db["users"].items()
        if ud.get("department")==dept and ud["role"] in ("faculty","dept_head")
    ]
    return jsonify({
        "timetable": tt,
        "time_slots": TIME_SLOTS,
        "days": DAYS,
        "sections": ALL_SECTIONS.get(dept,[]),
        "cameras": CAMERA_SOURCES,
        "faculty_list": faculty_list,
    })

@app.route("/dept/timetable_full")
def dept_timetable_full():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    faculty_username = request.args.get("faculty","")
    if not faculty_username:
        return jsonify({"error":"faculty required"}),400
    tt = db["timetable"].get(faculty_username, {})
    return jsonify({"timetable": tt, "time_slots": TIME_SLOTS, "days": DAYS})

@app.route("/dept/save_timetable", methods=["POST"])
def dept_save_timetable():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d = request.get_json(silent=True) or {}
    faculty_username = d.get("faculty_username","")
    day     = d.get("day","Monday")
    entries = d.get("entries",{})

    if not faculty_username:
        return jsonify({"error":"faculty_username required"}),400

    db["timetable"].setdefault(faculty_username, {})
    db["timetable"][faculty_username][day] = entries
    save_db()

    faculty_user = db["users"].get(faculty_username, {})
    dept = faculty_user.get("department", u.get("department",""))
    # Regenerate sessions for past 4 weeks, current week and next 2 weeks
    # so changes are immediately visible in both faculty calendar and dept sessions
    for offset in range(-4, 3):
        generate_sessions_for_week(faculty_username, dept, offset)

    return jsonify({"success": True})

@app.route("/dept/upload_timetable_excel", methods=["POST"])
def dept_upload_timetable_excel():
    """
    Bulk-import timetable for ALL faculty from a single Excel file.
    Expected columns (case-insensitive):
        Faculty - faculty username (required)
        Day     - Monday ... Saturday (required)
        Period  - slot label OR slot id e.g. 'Period 1'/'slot1' (required)
        Section - section code e.g. 'CSE-3' (optional)
        Subject - subject name (optional)
        Camera  - camera id (optional, default 'default')
    """
    u = me()
    if not u or u["role"] not in ("admin", "dept_head"):
        return jsonify({"error": "Forbidden"}), 403
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error": "openpyxl not installed"}), 500
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files accepted"}), 400
    slot_by_label = {s["label"].lower(): s["id"] for s in TIME_SLOTS}
    slot_by_id    = {s["id"]: s["id"] for s in TIME_SLOTS}
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        raw_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        def col(keyword):
            return next((i for i, h in enumerate(raw_headers) if keyword in h), None)
        fac_col  = col("faculty")
        day_col  = col("day")
        per_col  = next((i for i, h in enumerate(raw_headers) if "period" in h or "slot" in h), None)
        sec_col  = col("section")
        subj_col = col("subject")
        cam_col  = col("camera")
        if fac_col is None or day_col is None or per_col is None:
            return jsonify({"error": "Required columns missing. Need: Faculty, Day, Period (or Slot)"}), 400
        dept_for_check = u.get("department") if u["role"] == "dept_head" else None
        updated_faculty = set()
        skipped = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            fac_uname = str(row[fac_col]).strip() if row[fac_col] else ""
            day_val   = str(row[day_col]).strip().capitalize() if row[day_col] else ""
            period_raw= str(row[per_col]).strip() if row[per_col] else ""
            section   = str(row[sec_col]).strip() if sec_col is not None and row[sec_col] else ""
            subject   = str(row[subj_col]).strip() if subj_col is not None and row[subj_col] else ""
            cam_id    = str(row[cam_col]).strip() if cam_col is not None and row[cam_col] else "default"
            if not fac_uname or fac_uname.lower() == "none":
                continue
            if day_val not in DAYS:
                skipped.append(f"Row {row_idx}: unknown day '{day_val}'")
                continue
            slot_id = slot_by_id.get(period_raw.lower()) or slot_by_label.get(period_raw.lower())
            if not slot_id:
                skipped.append(f"Row {row_idx}: unknown period '{period_raw}'")
                continue
            if slot_id in BREAK_SLOTS:
                continue
            fac_user = db["users"].get(fac_uname)
            if not fac_user:
                skipped.append(f"Row {row_idx}: faculty '{fac_uname}' not found")
                continue
            if dept_for_check and fac_user.get("department") != dept_for_check:
                skipped.append(f"Row {row_idx}: faculty '{fac_uname}' not in your dept")
                continue
            db["timetable"].setdefault(fac_uname, {}).setdefault(day_val, {})
            db["timetable"][fac_uname][day_val][slot_id] = {
                "section": section, "subject": subject, "cam_id": cam_id,
            }
            updated_faculty.add(fac_uname)
        for fac_uname in updated_faculty:
            fac_user = db["users"].get(fac_uname, {})
            dept = fac_user.get("department", u.get("department", ""))
            for offset in range(-4, 3):
                generate_sessions_for_week(fac_uname, dept, offset)
        save_db()
        return jsonify({"success": True, "faculty_updated": len(updated_faculty), "skipped": skipped})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/dept/timetable_template_excel")
def dept_timetable_template_excel():
    """Download a polished, pre-filled template Excel for bulk timetable upload."""
    u = me()
    if not u or u["role"] not in ("admin", "dept_head"):
        return Response("Forbidden", status=403)
    if not OPENPYXL_AVAILABLE:
        return Response("openpyxl not installed", status=500)

    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    dept = u["department"] if u["role"] == "dept_head" else "CSE"
    faculty_list = [
        (un, ud["name"])
        for un, ud in db["users"].items()
        if ud.get("department") == dept and ud["role"] in ("faculty", "dept_head")
    ]
    period_slots = [s for s in TIME_SLOTS if s["id"] not in BREAK_SLOTS]

    # ── colour palette ────────────────────────────────────────────────────────
    C_HEADER_BG   = "1F3864"   # dark navy  – sheet main header
    C_HEADER_FG   = "FFFFFF"
    C_DAY_BG      = "2E75B6"   # blue       – day group header
    C_DAY_FG      = "FFFFFF"
    C_SUN_BG      = "FFF2CC"   # pale amber – Sunday rows
    C_SAT_BG      = "EBF3FB"   # pale blue  – Saturday rows
    C_ALT_BG      = "F5F8FF"   # very light – alternating data rows
    C_WHITE       = "FFFFFF"
    C_BORDER      = "B8CCE4"
    C_REQ_BG      = "FFF2CC"   # yellow     – required-field header
    C_OPT_BG      = "E2EFDA"   # green      – optional-field header

    def _font(bold=False, size=11, color="000000", italic=False):
        return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _border(style="thin", color=C_BORDER):
        s = Side(style=style, color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 – Timetable Data
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Timetable Data"
    ws.freeze_panes = "A4"          # freeze title + column-header rows
    ws.sheet_view.showGridLines = False

    # ── Row 1: big title banner ───────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"📅  {dept} Department — Timetable Upload Template"
    title_cell.font  = _font(bold=True, size=14, color=C_HEADER_FG)
    title_cell.fill  = _fill(C_HEADER_BG)
    title_cell.alignment = _align("center")
    ws.row_dimensions[1].height = 30

    # ── Row 2: subtitle / instructions line ───────────────────────────────────
    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value = ("Fill in Section and Subject for each period. "
                 "Leave the row blank to skip that period. "
                 "Yellow columns are required  •  Green columns are optional.")
    sub.font      = _font(size=9, color="595959", italic=True)
    sub.fill      = _fill("D9E1F2")
    sub.alignment = _align("center")
    ws.row_dimensions[2].height = 18

    # ── Row 3: column headers ─────────────────────────────────────────────────
    COL_HEADERS = [
        ("Faculty Username",  C_REQ_BG, 22),
        ("Faculty Name",      "DAEEF3", 22),
        ("Day",               C_REQ_BG, 14),
        ("Period",            C_REQ_BG, 14),
        ("Time",              "DAEEF3", 16),
        ("Section",           C_REQ_BG, 16),
        ("Subject",           C_REQ_BG, 24),
    ]
    for col_idx, (label, bg, width) in enumerate(COL_HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font      = _font(bold=True, size=10, color="1F3864")
        cell.fill      = _fill(bg)
        cell.alignment = _align("center")
        cell.border    = _border("medium", "2E75B6")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    row = 4
    day_colors = {
        "Sunday":   C_SUN_BG,
        "Saturday": C_SAT_BG,
    }

    # Data-validation lists (hidden sheet referenced below)
    fac_usernames = [f for f, _ in faculty_list] or ["faculty_username"]
    fac_name_map  = {f: n for f, n in faculty_list}
    day_list      = ",".join(DAYS)
    period_list   = ",".join(s["label"] for s in period_slots)

    dv_day = DataValidation(type="list", formula1=f'"{day_list}"',
                            showDropDown=False, showErrorMessage=True,
                            error="Choose a valid day", errorTitle="Invalid Day")
    dv_period = DataValidation(type="list", formula1=f'"{period_list}"',
                               showDropDown=False, showErrorMessage=True,
                               error="Choose a valid period", errorTitle="Invalid Period")
    ws.add_data_validation(dv_day)
    ws.add_data_validation(dv_period)

    all_data_rows = []
    for fac_un, fac_name in faculty_list:
        for day in DAYS:
            for slot in period_slots:
                all_data_rows.append((fac_un, fac_name, day, slot))

    # If no faculty yet, add sample rows
    if not all_data_rows:
        for day in DAYS:
            for slot in period_slots:
                all_data_rows.append(("faculty_username", "Faculty Full Name",
                                      day, slot))

    for fac_un, fac_name, day, slot in all_data_rows:
        time_str = f"{slot['start']} – {slot['end']}"
        row_data = [fac_un, fac_name, day, slot["label"], time_str, "", ""]

        bg = day_colors.get(day, C_WHITE if row % 2 == 0 else C_ALT_BG)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.fill      = _fill(bg)
            cell.font      = _font(size=10)
            cell.alignment = _align("center" if col_idx in (3,4,5) else "left")
            cell.border    = _border("thin")

        # Lock Faculty / Day / Period / Time as read-only hints (light grey text)
        for col_idx in (1, 2, 3, 4, 5):
            ws.cell(row=row, column=col_idx).font = _font(size=10, color="595959")

        dv_day.add(ws.cell(row=row, column=3))
        dv_period.add(ws.cell(row=row, column=4))
        row += 1

    ws.row_dimensions[row].height = 6   # small gap at end

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 – Faculty List  (quick reference)
    # ══════════════════════════════════════════════════════════════════════════
    wf = wb.create_sheet("Faculty List")
    wf.sheet_view.showGridLines = False

    wf.merge_cells("A1:C1")
    hdr = wf["A1"]
    hdr.value     = f"Faculty Reference — {dept}"
    hdr.font      = _font(bold=True, size=12, color=C_HEADER_FG)
    hdr.fill      = _fill(C_HEADER_BG)
    hdr.alignment = _align("center")
    wf.row_dimensions[1].height = 26

    for ci, (label, w) in enumerate(
            [("Username (use in template)", 28), ("Full Name", 28), ("Department", 18)],
            start=1):
        c = wf.cell(row=2, column=ci, value=label)
        c.font      = _font(bold=True, size=10, color="1F3864")
        c.fill      = _fill("D9E1F2")
        c.alignment = _align("center")
        c.border    = _border("medium", "2E75B6")
        wf.column_dimensions[get_column_letter(ci)].width = w
    wf.row_dimensions[2].height = 20

    for i, (un, name) in enumerate(faculty_list, start=3):
        bg = C_WHITE if i % 2 == 0 else C_ALT_BG
        for ci, val in enumerate([un, name, dept], start=1):
            c = wf.cell(row=i, column=ci, value=val)
            c.font      = _font(size=10)
            c.fill      = _fill(bg)
            c.alignment = _align("left")
            c.border    = _border()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 – Instructions
    # ══════════════════════════════════════════════════════════════════════════
    wi = wb.create_sheet("Instructions")
    wi.sheet_view.showGridLines = False
    wi.column_dimensions["A"].width = 26
    wi.column_dimensions["B"].width = 48
    wi.column_dimensions["C"].width = 42

    wi.merge_cells("A1:C1")
    t = wi["A1"]
    t.value     = "📖  How to Fill the Timetable Template"
    t.font      = _font(bold=True, size=13, color=C_HEADER_FG)
    t.fill      = _fill(C_HEADER_BG)
    t.alignment = _align("center")
    wi.row_dimensions[1].height = 28

    # column headers
    for ci, label in enumerate(["Column", "What to Enter", "Valid Values / Notes"], start=1):
        c = wi.cell(row=2, column=ci, value=label)
        c.font      = _font(bold=True, size=10, color="1F3864")
        c.fill      = _fill("D9E1F2")
        c.alignment = _align("center")
        c.border    = _border("medium", "2E75B6")
    wi.row_dimensions[2].height = 20

    instructions = [
        ("Faculty Username", "Login username of the faculty member",
         "See 'Faculty List' sheet for all valid usernames"),
        ("Faculty Name",     "Auto-filled for reference — do not change",
         "Read-only reference column"),
        ("Day",              "Day of the week for this class",
         f"{', '.join(DAYS)}\n(Sunday = extra / special classes only)"),
        ("Period",           "Period label for the class",
         ", ".join(s["label"] for s in period_slots)),
        ("Time",             "Auto-filled — do not change",
         "Read-only, shows period start–end time"),
        ("Section",          "Class section (REQUIRED)",
         "e.g.  CSE-A,  ECE-2,  MECH-B"),
        ("Subject",          "Subject / course name (REQUIRED)",
         "e.g.  Data Structures,  DBMS,  Thermodynamics"),
    ]

    for i, (col, what, vals) in enumerate(instructions, start=3):
        bg = C_WHITE if i % 2 == 0 else C_ALT_BG
        data = [(col, "1F3864", True), (what, "000000", False), (vals, "595959", False)]
        for ci, (val, clr, bold) in enumerate(data, start=1):
            c = wi.cell(row=i, column=ci, value=val)
            c.font      = _font(bold=bold, size=10, color=clr)
            c.fill      = _fill(bg)
            c.alignment = _align("left", wrap=True)
            c.border    = _border()
        wi.row_dimensions[i].height = 30

    # Tips section
    tip_row = len(instructions) + 4
    wi.merge_cells(f"A{tip_row}:C{tip_row}")
    t2 = wi.cell(row=tip_row, column=1, value="💡  Tips")
    t2.font      = _font(bold=True, size=11, color=C_HEADER_FG)
    t2.fill      = _fill(C_DAY_BG)
    t2.alignment = _align("left")
    wi.row_dimensions[tip_row].height = 22

    tips = [
        "Leave Section and Subject empty to skip a period — the row will be ignored on upload.",
        "Sunday rows are included for extra/special classes only. Leave them blank if unused.",
        "You can delete entire rows for periods you never use to keep the file tidy.",
        "Do NOT change Faculty Username, Day, Period, or Time columns — they are pre-filled.",
        "Upload this file via Department → Timetable → Upload Excel Timetable.",
    ]
    for j, tip in enumerate(tips, start=tip_row + 1):
        wi.merge_cells(f"A{j}:C{j}")
        c = wi.cell(row=j, column=1, value=f"  ✔  {tip}")
        c.font      = _font(size=10, color="1F3864")
        c.fill      = _fill(C_ALT_BG if j % 2 == 0 else C_WHITE)
        c.alignment = _align("left", wrap=True)
        c.border    = _border()
        wi.row_dimensions[j].height = 18

    # ── finalise ──────────────────────────────────────────────────────────────
    # Set active sheet back to data sheet
    wb.active = ws

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f"attachment; filename=timetable_template_{dept}.xlsx"}
    )


# ── Sessions ──────────────────────────────────────────────────────────────────
@app.route("/dept/sessions")
def dept_sessions():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    dept = u["department"] if u["role"]=="dept_head" else request.args.get("dept","CSE")
    section = request.args.get("section","")
    date_filter = request.args.get("date","")
    week_offset = int(request.args.get("week", 0))

    # Generate sessions for every faculty in this dept for the requested week
    # so faculty calendar is always in sync with what dept_head sees
    for uname, ud in db["users"].items():
        if ud.get("department") == dept and ud["role"] in ("faculty","dept_head"):
            generate_sessions_for_week(uname, dept, week_offset)

    # Compute week date range for filtering when no specific date_filter provided
    today = datetime.date.today()
    monday = today - datetime.timedelta(days=today.weekday()) + datetime.timedelta(weeks=week_offset)
    sunday = monday + datetime.timedelta(days=6)

    sessions = []
    for s in db["sessions"]:
        if s.get("dept") != dept:
            continue
        if section and s.get("section") != section:
            continue
        if date_filter:
            if s.get("date") != date_filter:
                continue
        else:
            # Filter by week range
            try:
                sess_date = datetime.date.fromisoformat(s["date"])
            except Exception:
                continue
            if not (monday <= sess_date <= sunday):
                continue
        sessions.append(s)

    for s in sessions:
        key = s.get("att_key","")
        recs = db["attendance_records"].get(key, [])
        total = len(db["students"].get(s["section"], []))
        present = sum(1 for r in recs if r["status"] == "Present")
        s["att_count"] = present
        s["total_students"] = total
        s["att_marked"] = len(recs) >= total > 0
        orig_user = db["users"].get(s.get("faculty_username",""), {})
        s["faculty_name"] = orig_user.get("name", s.get("faculty_username",""))
        if s.get("substitute"):
            sub_user = db["users"].get(s["substitute"], {})
            s["substitute_name"] = sub_user.get("name", s["substitute"])
        else:
            s["substitute_name"] = None
    return jsonify({"sessions": sessions})

@app.route("/dept/assign_substitute", methods=["POST"])
def dept_assign_substitute():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d = request.get_json(silent=True) or {}
    session_id  = d.get("session_id","")
    substitute  = d.get("substitute","")

    sess = next((s for s in db["sessions"] if s["id"] == session_id), None)
    if not sess:
        return jsonify({"error":"Session not found"}),404

    sess["substitute"] = substitute
    save_db()

    # Regenerate sessions for BOTH the original faculty and the substitute
    # so the substitute's calendar immediately reflects the new assignment
    for uname in {sess["faculty_username"], substitute}:
        if not uname:
            continue
        fac_user = db["users"].get(uname, {})
        dept = fac_user.get("department", sess.get("dept", ""))
        for offset in range(-4, 3):
            generate_sessions_for_week(uname, dept, offset)

    return jsonify({"success": True})

# ── Analytics ─────────────────────────────────────────────────────────────────
@app.route("/dept/analytics")
def dept_analytics():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    dept    = u["department"] if u["role"]=="dept_head" else request.args.get("dept","CSE")
    section = request.args.get("section","")
    date    = request.args.get("date", datetime.date.today().isoformat())

    total_stu = len(db["students"].get(section,[]))
    summary = []
    for slot in TIME_SLOTS:
        if slot["id"] in BREAK_SLOTS:
            continue
        key = f"{section}::{slot['id']}::{date}"
        recs = db["attendance_records"].get(key, [])
        present = sum(1 for r in recs if r.get("status") == "Present")
        absent  = max(0, total_stu - present)
        summary.append({
            "slot_id":   slot["id"],
            "label":     slot["label"],
            "time":      f"{slot['start']}–{slot['end']}",
            "present":   present,
            "absent":    absent,
            "total":     total_stu,
            "pct":       round(present/total_stu*100,1) if total_stu else 0,
            "records":   recs,
        })
    return jsonify({"summary": summary, "section": section, "date": date, "dept": dept})

@app.route("/dept/attendance_section")
def dept_attendance_section():
    u = me()
    if not u or u["role"] not in ("admin","dept_head","faculty"):
        return jsonify({"error":"Forbidden"}),403
    section  = request.args.get("section","")
    slot_id  = request.args.get("slot","")
    date_str = request.args.get("date", datetime.date.today().isoformat())
    key = f"{section}::{slot_id}::{date_str}"
    recs = db["attendance_records"].get(key, [])
    all_students = db["students"].get(section, [])
    present_rolls = {r["roll"] for r in recs if r["status"] == "Present"}
    result = []
    for stu in all_students:
        result.append({
            "roll":   stu["roll"],
            "name":   stu["name"],
            "status": "Present" if stu["roll"] in present_rolls else "Absent",
            "photo":  get_student_photo_b64(stu["roll"]),
        })
    return jsonify({"records": result, "section": section, "slot_id": slot_id, "date": date_str})

@app.route("/dept/update_attendance", methods=["POST"])
def dept_update_attendance():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d        = request.get_json(silent=True) or {}
    section  = d.get("section","")
    slot_id  = d.get("slot_id","")
    date_str = d.get("date", datetime.date.today().isoformat())
    records  = d.get("records",[])

    try:
        att_date = datetime.date.fromisoformat(date_str)
        cutoff   = datetime.datetime.combine(att_date + datetime.timedelta(days=1),
                                              datetime.time(18, 0))
        if datetime.datetime.now() > cutoff:
            return jsonify({"error": "Attendance can only be modified until 6 PM the next day"}), 403
    except Exception:
        pass

    key = f"{section}::{slot_id}::{date_str}"
    # Normalize records to ensure consistent structure
    clean_records = []
    for r in records:
        if not r.get("roll"):
            continue
        clean_records.append({
            "roll":      r["roll"],
            "name":      r.get("name", r["roll"]),
            "status":    r.get("status", "Absent"),
            "time":      datetime.datetime.now().strftime("%H:%M:%S"),
            "marked_by": session.get("username", "dept_head"),
        })
    db["attendance_records"][key] = clean_records
    save_db()
    dept = u.get("department", "")
    rewrite_attendance_csv_for_key(section, dept, slot_id, date_str, clean_records, u.get("name",""))
    return jsonify({"success": True})

# ── Student search ────────────────────────────────────────────────────────────
@app.route("/dept/search_attendance")
def dept_search_attendance():
    u = me()
    if not u or u["role"] not in ("admin","dept_head","faculty"):
        return jsonify({"error":"Forbidden"}),403
    query   = request.args.get("q","").strip().lower()
    section = request.args.get("section","")
    dept    = u["department"] if u["role"] in ("dept_head","faculty") else request.args.get("dept","")

    results = {}
    for key, recs in db["attendance_records"].items():
        parts = key.split("::")
        if len(parts) != 3:
            continue
        sec, slot_id, date_str = parts
        if section and sec != section:
            continue
        for r in recs:
            roll = r.get("roll","").lower()
            name = r.get("name","").lower()
            if query and query not in roll and query not in name:
                continue
            results.setdefault(r["roll"], {
                "roll": r["roll"],
                "name": r.get("name",""),
                "photo": get_student_photo_b64(r["roll"]),
                "records": []
            })
            results[r["roll"]]["records"].append({
                "section": sec,
                "slot_id": slot_id,
                "date": date_str,
                "status": r.get("status","Present"),
            })

    return jsonify({"results": list(results.values())})

# ── Excel upload ──────────────────────────────────────────────────────────────
@app.route("/dept/upload_students_excel", methods=["POST"])
def dept_upload_students_excel():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error":"openpyxl not installed"}),500

    file    = request.files.get("file")
    section = request.form.get("section","")

    if not file:
        return jsonify({"error":"No file uploaded"}),400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error":"Only .xlsx files accepted"}),400

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        roll_col = next((i for i,h in enumerate(headers) if "roll" in h), None)
        name_col = next((i for i,h in enumerate(headers) if "name" in h), None)
        if roll_col is None:
            return jsonify({"error":"No 'Roll' column found in Excel"}),400

        students = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            roll = str(row[roll_col]).strip() if row[roll_col] else ""
            name = str(row[name_col]).strip() if (name_col is not None and row[name_col]) else roll
            if not roll or roll.lower() == "none":
                continue
            students.append({"roll": roll, "name": name})

        db["students"][section] = students
        save_db()
        return jsonify({"success": True, "count": len(students)})
    except Exception as e:
        return jsonify({"error": str(e)}),500

@app.route("/dept/upload_excel", methods=["POST"])
def dept_upload_excel():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error":"openpyxl not installed"}),500

    file    = request.files.get("file")
    section = request.form.get("section","")
    slot_id = request.form.get("slot_id","")
    date_str = request.form.get("date", datetime.date.today().isoformat())
    dept    = u["department"] if u["role"]=="dept_head" else request.form.get("dept","")

    if not file:
        return jsonify({"error":"No file uploaded"}),400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error":"Only .xlsx files accepted"}),400

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        roll_col   = next((i for i,h in enumerate(headers) if "roll" in h), None)
        name_col   = next((i for i,h in enumerate(headers) if "name" in h), None)
        status_col = next((i for i,h in enumerate(headers) if "status" in h), None)
        if roll_col is None:
            return jsonify({"error":"No 'Roll' column found"}),400

        stu_map = {s["roll"]: s["name"] for s in db["students"].get(section,[])}
        key = f"{section}::{slot_id}::{date_str}"
        existing_rolls = {r["roll"] for r in db["attendance_records"].get(key,[])}
        db["attendance_records"].setdefault(key, [])
        count = 0
        now = datetime.datetime.now().strftime("%H:%M:%S")
        for row in ws.iter_rows(min_row=2, values_only=True):
            roll = str(row[roll_col]).strip() if row[roll_col] else ""
            if not roll or roll.lower() == "none":
                continue
            name   = str(row[name_col]).strip() if (name_col is not None and row[name_col]) else stu_map.get(roll, roll)
            status = str(row[status_col]).strip() if (status_col is not None and row[status_col]) else "Present"
            if roll not in existing_rolls:
                db["attendance_records"][key].append({"roll":roll,"name":name,"status":status,"time":now,"marked_by":"excel"})
                existing_rolls.add(roll)
                write_attendance_csv(roll, name, section, dept, slot_id, "", u.get("name",""), status)
                count += 1

        save_db()
        return jsonify({"success": True, "marked": count})
    except Exception as e:
        return jsonify({"error": str(e)}),500

# ── Students CRUD ─────────────────────────────────────────────────────────────
@app.route("/dept/students")
def dept_students():
    u = me()
    if not u or u["role"] not in ("admin","dept_head","faculty"):
        return jsonify({"error":"Forbidden"}),403
    section = request.args.get("section","")
    students = db["students"].get(section,[])
    result = []
    for s in students:
        result.append({**s, "photo": get_student_photo_b64(s["roll"])})
    return jsonify({"students": result})

@app.route("/dept/save_students", methods=["POST"])
def dept_save_students():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d = request.get_json(silent=True) or {}
    section  = d.get("section","")
    students = d.get("students",[])
    db["students"][section] = students
    save_db()
    return jsonify({"success": True})

@app.route("/dept/add_student", methods=["POST"])
def dept_add_student():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d = request.get_json(silent=True) or {}
    section = d.get("section","")
    roll    = d.get("roll","").strip()
    name    = d.get("name","").strip()
    if not section or not roll:
        return jsonify({"error":"section and roll required"}),400
    students = db["students"].setdefault(section, [])
    if any(s["roll"] == roll for s in students):
        return jsonify({"error":"Roll number already exists in this section"}),409
    students.append({"roll": roll, "name": name or roll})
    save_db()
    return jsonify({"success": True})

@app.route("/dept/remove_student", methods=["POST"])
def dept_remove_student():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d = request.get_json(silent=True) or {}
    section = d.get("section","")
    roll    = d.get("roll","")
    if section in db["students"]:
        db["students"][section] = [s for s in db["students"][section] if s["roll"] != roll]
        save_db()
    return jsonify({"success": True})

# ── Download attendance ───────────────────────────────────────────────────────
@app.route("/dept/download_attendance")
def dept_download_attendance():
    u = me()
    if not u or u["role"] not in ("admin","dept_head","faculty"):
        return Response("Forbidden",status=403)
    if not OPENPYXL_AVAILABLE:
        return Response("openpyxl not installed",status=500)
    import openpyxl as _openpyxl

    section = request.args.get("section","")
    date    = request.args.get("date", datetime.date.today().isoformat())

    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{section} {date}"
    ws.append(["Roll", "Name", "Period 1", "Period 2", "Period 3",
               "Period 4", "Period 5", "Period 6", "Period 7", "Period 8"])

    all_students = db["students"].get(section, [])
    period_slots = [s for s in TIME_SLOTS if s["id"] not in BREAK_SLOTS]

    for stu in all_students:
        row = [stu["roll"], stu["name"]]
        for slot in period_slots:
            key = f"{section}::{slot['id']}::{date}"
            recs = db["attendance_records"].get(key, [])
            present_rolls = {r["roll"] for r in recs if r["status"] == "Present"}
            row.append("P" if stu["roll"] in present_rolls else "A")
        ws.append(row)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"attendance_{section}_{date}.xlsx"
    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

@app.route("/dept/download_student_attendance")
def dept_download_student_attendance():
    u = me()
    if not u or u["role"] not in ("admin","dept_head","faculty"):
        return Response("Forbidden",status=403)
    if not OPENPYXL_AVAILABLE:
        return Response("openpyxl not installed",status=500)

    query   = request.args.get("q","").strip().lower()
    section = request.args.get("section","")

    results = {}
    for key, recs in db["attendance_records"].items():
        parts = key.split("::")
        if len(parts) != 3:
            continue
        sec, slot_id, date_str = parts
        if section and sec != section:
            continue
        slot_label = next((s["label"] for s in TIME_SLOTS if s["id"]==slot_id), slot_id)
        for r in recs:
            roll = r.get("roll","").lower()
            name = r.get("name","").lower()
            if query and query not in roll and query not in name:
                continue
            results.setdefault(r["roll"], {
                "roll": r["roll"], "name": r.get("name",""), "records": []
            })
            results[r["roll"]]["records"].append({
                "section": sec, "slot_id": slot_id, "slot_label": slot_label,
                "date": date_str, "status": r.get("status","Present"),
            })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Attendance"
    ws.append(["Roll", "Name", "Section", "Date", "Period", "Status"])
    for stu in results.values():
        for rec in sorted(stu["records"], key=lambda x: (x["date"],x["slot_id"])):
            ws.append([stu["roll"], stu["name"], rec["section"], rec["date"], rec["slot_label"], rec["status"]])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"student_attendance_{query or 'all'}.xlsx"
    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

# ── Admin routes ──────────────────────────────────────────────────────────────
@app.route("/admin/all_users")
def admin_all_users():
    u = me()
    if not u or u["role"] != "admin":
        return jsonify({"error":"Forbidden"}),403
    users = [
        {"username": un, **{k:v for k,v in ud.items() if k!="password_hash"}}
        for un, ud in db["users"].items()
    ]
    return jsonify({"users": users, "departments": DEPARTMENTS})

@app.route("/admin/export_csv")
def admin_export_csv():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return Response("Forbidden",status=403)
    section = request.args.get("section","")
    dept    = u["department"] if u["role"]=="dept_head" else request.args.get("dept","")

    out_rows = []
    if os.path.exists(CSV_FILE):
        with open(CSV_FILE) as f:
            reader = csv.DictReader(f)
            for row in reader:
                if section and row.get("Section")!=section:
                    continue
                if dept and row.get("Department")!=dept:
                    continue
                out_rows.append(row)

    out = io.StringIO()
    if out_rows:
        w = csv.DictWriter(out, fieldnames=out_rows[0].keys())
        w.writeheader(); w.writerows(out_rows)
    fname = f"attendance_{section or dept or 'all'}.csv"
    return Response(out.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})

@app.route("/")
def index():
    try:
        with open("index.html") as f:
            return f.read(), 200, {"Content-Type": "text/html"}
    except FileNotFoundError:
        return "<h2>Put index.html alongside app.py</h2>", 200

# ═══════════════════════════════════════════════════════════════════════════════
#  STARTUP
# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    setup_csv()
    init_face_model()
    load_known_faces()
    for cam in CAMERA_SOURCES:
        threading.Thread(target=camera_loop, args=(cam["id"],cam["source"]), daemon=True).start()
    print("\n🌐  http://localhost:8080")
    print("🔑  Default admin login: admin / admin123\n")
    app.run(host="0.0.0.0", port=8080, debug=False, threaded=True)
