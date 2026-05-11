from flask import Blueprint, request, jsonify, send_file
import datetime, io, base64
from database import db, _state_lock, save_db, rewrite_attendance_csv_for_key, _hash
from auth import me
from config import DAYS, ALL_SECTIONS, DEPARTMENTS, TIME_SLOTS
from session_engine import get_faculty_sessions

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

dept_bp = Blueprint('dept', __name__)


def get_student_photo_b64(roll: str):
    for ext in (".jpg", ".jpeg", ".png"):
        path = os.path.join(KNOWN_FACES_DIR, roll + ext)
        if os.path.exists(path):
            try:
                with open(path, "rb") as f:
                    data = base64.b64encode(f.read()).decode()
                mime = "image/jpeg" if ext in (".jpg", ".jpeg") else "image/png"
                return f"data:{mime};base64,{data}"
            except Exception:
                pass
    return None

@dept_bp.route("/dept/faculty_list")
def dept_faculty_list():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    dept = u["department"] if u["role"]=="dept_head" else request.args.get("dept","CSE")
    users = [
        {"username": un, "name": ud["name"], "role": ud["role"],
         "sections": ud.get("sections",[]), "department": ud.get("department")}
        for un, ud in db["users"].items()
        if ud.get("department") == dept and ud["role"] in ("faculty","dept_head")
    ]
    return jsonify({"faculty": users, "sections": ALL_SECTIONS.get(dept,[])})

@dept_bp.route("/dept/create_faculty", methods=["POST"])
def dept_create_faculty():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d = request.get_json(silent=True) or {}
    uname    = d.get("username","").strip()
    password = d.get("password","")
    name     = d.get("name","")
    role     = d.get("role","faculty")
    dept     = d.get("department","")
    sections = d.get("sections",[])

    if u["role"]=="dept_head":
        dept = u["department"]
        if role == "admin":
            return jsonify({"error":"Cannot create admin"}),403
    if not uname or not password:
        return jsonify({"error":"Username and password required"}),400
    if uname in db["users"]:
        return jsonify({"error":"Username already exists"}),409

    db["users"][uname] = {
        "password_hash": _hash(password),
        "role": role,
        "department": dept,
        "name": name or uname,
        "sections": sections,
    }
    save_db()
    return jsonify({"success": True})

@dept_bp.route("/dept/update_faculty", methods=["POST"])
def dept_update_faculty():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d = request.get_json(silent=True) or {}
    uname = d.get("username","")
    if uname not in db["users"]:
        return jsonify({"error":"User not found"}),404
    target = db["users"][uname]
    if u["role"]=="dept_head" and target.get("department")!=u["department"]:
        return jsonify({"error":"Cross-department forbidden"}),403
    if "name"     in d: target["name"]     = d["name"]
    if "sections" in d: target["sections"] = d["sections"]
    if "role"     in d and u["role"]=="admin": target["role"] = d["role"]
    if "password" in d and d["password"]: target["password_hash"] = _hash(d["password"])
    save_db()
    return jsonify({"success": True})

@dept_bp.route("/dept/delete_faculty", methods=["POST"])
def dept_delete_faculty():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d = request.get_json(silent=True) or {}
    uname = d.get("username","")
    if uname == "admin":
        return jsonify({"error":"Cannot delete root admin"}),403
    if uname not in db["users"]:
        return jsonify({"error":"Not found"}),404
    del db["users"][uname]
    save_db()
    return jsonify({"success": True})

@dept_bp.route("/dept/timetable")
def dept_timetable():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    dept = u["department"] if u["role"]=="dept_head" else request.args.get("dept","CSE")
    faculty_username = request.args.get("faculty","")
    day = request.args.get("day","Monday")
    tt = db["timetable"].get(faculty_username, {}).get(day, {}) if faculty_username else {}
    faculty_list = [
        {"username": un, "name": ud["name"]}
        for un, ud in db["users"].items()
        if ud.get("department")==dept and ud["role"] in ("faculty","dept_head")
    ]
    return jsonify({
        "timetable": tt,
        "time_slots": TIME_SLOTS,
        "days": DAYS,
        "sections": ALL_SECTIONS.get(dept,[]),
        "cameras": CAMERA_SOURCES,
        "faculty_list": faculty_list,
    })

@dept_bp.route("/dept/timetable_full")
def dept_timetable_full():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    faculty_username = request.args.get("faculty","")
    if not faculty_username:
        return jsonify({"error":"faculty required"}),400
    tt = db["timetable"].get(faculty_username, {})
    return jsonify({"timetable": tt, "time_slots": TIME_SLOTS, "days": DAYS})

@dept_bp.route("/dept/save_timetable", methods=["POST"])
def dept_save_timetable():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d = request.get_json(silent=True) or {}
    faculty_username = d.get("faculty_username","")
    day     = d.get("day","Monday")
    entries = d.get("entries",{})

    if not faculty_username:
        return jsonify({"error":"faculty_username required"}),400

    db["timetable"].setdefault(faculty_username, {})
    db["timetable"][faculty_username][day] = entries
    save_db()

    faculty_user = db["users"].get(faculty_username, {})
    dept = faculty_user.get("department", u.get("department",""))
    # Regenerate sessions for past 4 weeks, current week and next 2 weeks
    # so changes are immediately visible in both faculty calendar and dept sessions
    for offset in range(-4, 3):
        generate_sessions_for_week(faculty_username, dept, offset)

    return jsonify({"success": True})

@dept_bp.route("/dept/upload_timetable_excel", methods=["POST"])
def dept_upload_timetable_excel():
    """
    Bulk-import timetable for ALL faculty from a single Excel file.
    Expected columns (case-insensitive):
        Faculty - faculty username (required)
        Day     - Monday ... Saturday (required)
        Period  - slot label OR slot id e.g. 'Period 1'/'slot1' (required)
        Section - section code e.g. 'CSE-3' (optional)
        Subject - subject name (optional)
        Camera  - camera id (optional, default 'default')
    """
    u = me()
    if not u or u["role"] not in ("admin", "dept_head"):
        return jsonify({"error": "Forbidden"}), 403
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error": "openpyxl not installed"}), 500
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files accepted"}), 400
    slot_by_label = {s["label"].lower(): s["id"] for s in TIME_SLOTS}
    slot_by_id    = {s["id"]: s["id"] for s in TIME_SLOTS}
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        raw_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        def col(keyword):
            return next((i for i, h in enumerate(raw_headers) if keyword in h), None)
        fac_col  = col("faculty")
        day_col  = col("day")
        per_col  = next((i for i, h in enumerate(raw_headers) if "period" in h or "slot" in h), None)
        sec_col  = col("section")
        subj_col = col("subject")
        cam_col  = col("camera")
        if fac_col is None or day_col is None or per_col is None:
            return jsonify({"error": "Required columns missing. Need: Faculty, Day, Period (or Slot)"}), 400
        dept_for_check = u.get("department") if u["role"] == "dept_head" else None
        updated_faculty = set()
        skipped = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            fac_uname = str(row[fac_col]).strip() if row[fac_col] else ""
            day_val   = str(row[day_col]).strip().capitalize() if row[day_col] else ""
            period_raw= str(row[per_col]).strip() if row[per_col] else ""
            section   = str(row[sec_col]).strip() if sec_col is not None and row[sec_col] else ""
            subject   = str(row[subj_col]).strip() if subj_col is not None and row[subj_col] else ""
            cam_id    = str(row[cam_col]).strip() if cam_col is not None and row[cam_col] else "default"
            if not fac_uname or fac_uname.lower() == "none":
                continue
            if day_val not in DAYS:
                skipped.append(f"Row {row_idx}: unknown day '{day_val}'")
                continue
            slot_id = slot_by_id.get(period_raw.lower()) or slot_by_label.get(period_raw.lower())
            if not slot_id:
                skipped.append(f"Row {row_idx}: unknown period '{period_raw}'")
                continue
            if slot_id in BREAK_SLOTS:
                continue
            fac_user = db["users"].get(fac_uname)
            if not fac_user:
                skipped.append(f"Row {row_idx}: faculty '{fac_uname}' not found")
                continue
            if dept_for_check and fac_user.get("department") != dept_for_check:
                skipped.append(f"Row {row_idx}: faculty '{fac_uname}' not in your dept")
                continue
            db["timetable"].setdefault(fac_uname, {}).setdefault(day_val, {})
            db["timetable"][fac_uname][day_val][slot_id] = {
                "section": section, "subject": subject, "cam_id": cam_id,
            }
            updated_faculty.add(fac_uname)
        for fac_uname in updated_faculty:
            fac_user = db["users"].get(fac_uname, {})
            dept = fac_user.get("department", u.get("department", ""))
            for offset in range(-4, 3):
                generate_sessions_for_week(fac_uname, dept, offset)
        save_db()
        return jsonify({"success": True, "faculty_updated": len(updated_faculty), "skipped": skipped})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@dept_bp.route("/dept/timetable_template_excel")
def dept_timetable_template_excel():
    """Download a polished, pre-filled template Excel for bulk timetable upload."""
    u = me()
    if not u or u["role"] not in ("admin", "dept_head"):
        return Response("Forbidden", status=403)
    if not OPENPYXL_AVAILABLE:
        return Response("openpyxl not installed", status=500)

    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    dept = u["department"] if u["role"] == "dept_head" else "CSE"
    faculty_list = [
        (un, ud["name"])
        for un, ud in db["users"].items()
        if ud.get("department") == dept and ud["role"] in ("faculty", "dept_head")
    ]
    period_slots = [s for s in TIME_SLOTS if s["id"] not in BREAK_SLOTS]

    # ── colour palette ────────────────────────────────────────────────────────
    C_HEADER_BG   = "1F3864"   # dark navy  – sheet main header
    C_HEADER_FG   = "FFFFFF"
    C_DAY_BG      = "2E75B6"   # blue       – day group header
    C_DAY_FG      = "FFFFFF"
    C_SUN_BG      = "FFF2CC"   # pale amber – Sunday rows
    C_SAT_BG      = "EBF3FB"   # pale blue  – Saturday rows
    C_ALT_BG      = "F5F8FF"   # very light – alternating data rows
    C_WHITE       = "FFFFFF"
    C_BORDER      = "B8CCE4"
    C_REQ_BG      = "FFF2CC"   # yellow     – required-field header
    C_OPT_BG      = "E2EFDA"   # green      – optional-field header

    def _font(bold=False, size=11, color="000000", italic=False):
        return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _border(style="thin", color=C_BORDER):
        s = Side(style=style, color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 – Timetable Data
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Timetable Data"
    ws.freeze_panes = "A4"          # freeze title + column-header rows
    ws.sheet_view.showGridLines = False

    # ── Row 1: big title banner ───────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"📅  {dept} Department — Timetable Upload Template"
    title_cell.font  = _font(bold=True, size=14, color=C_HEADER_FG)
    title_cell.fill  = _fill(C_HEADER_BG)
    title_cell.alignment = _align("center")
    ws.row_dimensions[1].height = 30

    # ── Row 2: subtitle / instructions line ───────────────────────────────────
    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value = ("Fill in Section and Subject for each period. "
                 "Leave the row blank to skip that period. "
                 "Yellow columns are required  •  Green columns are optional.")
    sub.font      = _font(size=9, color="595959", italic=True)
    sub.fill      = _fill("D9E1F2")
    sub.alignment = _align("center")
    ws.row_dimensions[2].height = 18

    # ── Row 3: column headers ─────────────────────────────────────────────────
    COL_HEADERS = [
        ("Faculty Username",  C_REQ_BG, 22),
        ("Faculty Name",      "DAEEF3", 22),
        ("Day",               C_REQ_BG, 14),
        ("Period",            C_REQ_BG, 14),
        ("Time",              "DAEEF3", 16),
        ("Section",           C_REQ_BG, 16),
        ("Subject",           C_REQ_BG, 24),
    ]
    for col_idx, (label, bg, width) in enumerate(COL_HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font      = _font(bold=True, size=10, color="1F3864")
        cell.fill      = _fill(bg)
        cell.alignment = _align("center")
        cell.border    = _border("medium", "2E75B6")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    row = 4
    day_colors = {
        "Sunday":   C_SUN_BG,
        "Saturday": C_SAT_BG,
    }

    # Data-validation lists (hidden sheet referenced below)
    fac_usernames = [f for f, _ in faculty_list] or ["faculty_username"]
    fac_name_map  = {f: n for f, n in faculty_list}
    day_list      = ",".join(DAYS)
    period_list   = ",".join(s["label"] for s in period_slots)

    dv_day = DataValidation(type="list", formula1=f'"{day_list}"',
                            showDropDown=False, showErrorMessage=True,
                            error="Choose a valid day", errorTitle="Invalid Day")
    dv_period = DataValidation(type="list", formula1=f'"{period_list}"',
                               showDropDown=False, showErrorMessage=True,
                               error="Choose a valid period", errorTitle="Invalid Period")
    ws.add_data_validation(dv_day)
    ws.add_data_validation(dv_period)

    all_data_rows = []
    for fac_un, fac_name in faculty_list:
        for day in DAYS:
            for slot in period_slots:
                all_data_rows.append((fac_un, fac_name, day, slot))

    # If no faculty yet, add sample rows
    if not all_data_rows:
        for day in DAYS:
            for slot in period_slots:
                all_data_rows.append(("faculty_username", "Faculty Full Name",
                                      day, slot))

    for fac_un, fac_name, day, slot in all_data_rows:
        time_str = f"{slot['start']} – {slot['end']}"
        row_data = [fac_un, fac_name, day, slot["label"], time_str, "", ""]

        bg = day_colors.get(day, C_WHITE if row % 2 == 0 else C_ALT_BG)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.fill      = _fill(bg)
            cell.font      = _font(size=10)
            cell.alignment = _align("center" if col_idx in (3,4,5) else "left")
            cell.border    = _border("thin")

        # Lock Faculty / Day / Period / Time as read-only hints (light grey text)
        for col_idx in (1, 2, 3, 4, 5):
            ws.cell(row=row, column=col_idx).font = _font(size=10, color="595959")

        dv_day.add(ws.cell(row=row, column=3))
        dv_period.add(ws.cell(row=row, column=4))
        row += 1

    ws.row_dimensions[row].height = 6   # small gap at end

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 – Faculty List  (quick reference)
    # ══════════════════════════════════════════════════════════════════════════
    wf = wb.create_sheet("Faculty List")
    wf.sheet_view.showGridLines = False

    wf.merge_cells("A1:C1")
    hdr = wf["A1"]
    hdr.value     = f"Faculty Reference — {dept}"
    hdr.font      = _font(bold=True, size=12, color=C_HEADER_FG)
    hdr.fill      = _fill(C_HEADER_BG)
    hdr.alignment = _align("center")
    wf.row_dimensions[1].height = 26

    for ci, (label, w) in enumerate(
            [("Username (use in template)", 28), ("Full Name", 28), ("Department", 18)],
            start=1):
        c = wf.cell(row=2, column=ci, value=label)
        c.font      = _font(bold=True, size=10, color="1F3864")
        c.fill      = _fill("D9E1F2")
        c.alignment = _align("center")
        c.border    = _border("medium", "2E75B6")
        wf.column_dimensions[get_column_letter(ci)].width = w
    wf.row_dimensions[2].height = 20

    for i, (un, name) in enumerate(faculty_list, start=3):
        bg = C_WHITE if i % 2 == 0 else C_ALT_BG
        for ci, val in enumerate([un, name, dept], start=1):
            c = wf.cell(row=i, column=ci, value=val)
            c.font      = _font(size=10)
            c.fill      = _fill(bg)
            c.alignment = _align("left")
            c.border    = _border()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 – Instructions
    # ══════════════════════════════════════════════════════════════════════════
    wi = wb.create_sheet("Instructions")
    wi.sheet_view.showGridLines = False
    wi.column_dimensions["A"].width = 26
    wi.column_dimensions["B"].width = 48
    wi.column_dimensions["C"].width = 42

    wi.merge_cells("A1:C1")
    t = wi["A1"]
    t.value     = "📖  How to Fill the Timetable Template"
    t.font      = _font(bold=True, size=13, color=C_HEADER_FG)
    t.fill      = _fill(C_HEADER_BG)
    t.alignment = _align("center")
    wi.row_dimensions[1].height = 28

    # column headers
    for ci, label in enumerate(["Column", "What to Enter", "Valid Values / Notes"], start=1):
        c = wi.cell(row=2, column=ci, value=label)
        c.font      = _font(bold=True, size=10, color="1F3864")
        c.fill      = _fill("D9E1F2")
        c.alignment = _align("center")
        c.border    = _border("medium", "2E75B6")
    wi.row_dimensions[2].height = 20

    instructions = [
        ("Faculty Username", "Login username of the faculty member",
         "See 'Faculty List' sheet for all valid usernames"),
        ("Faculty Name",     "Auto-filled for reference — do not change",
         "Read-only reference column"),
        ("Day",              "Day of the week for this class",
         f"{', '.join(DAYS)}\n(Sunday = extra / special classes only)"),
        ("Period",           "Period label for the class",
         ", ".join(s["label"] for s in period_slots)),
        ("Time",             "Auto-filled — do not change",
         "Read-only, shows period start–end time"),
        ("Section",          "Class section (REQUIRED)",
         "e.g.  CSE-A,  ECE-2,  MECH-B"),
        ("Subject",          "Subject / course name (REQUIRED)",
         "e.g.  Data Structures,  DBMS,  Thermodynamics"),
    ]

    for i, (col, what, vals) in enumerate(instructions, start=3):
        bg = C_WHITE if i % 2 == 0 else C_ALT_BG
        data = [(col, "1F3864", True), (what, "000000", False), (vals, "595959", False)]
        for ci, (val, clr, bold) in enumerate(data, start=1):
            c = wi.cell(row=i, column=ci, value=val)
            c.font      = _font(bold=bold, size=10, color=clr)
            c.fill      = _fill(bg)
            c.alignment = _align("left", wrap=True)
            c.border    = _border()
        wi.row_dimensions[i].height = 30

    # Tips section
    tip_row = len(instructions) + 4
    wi.merge_cells(f"A{tip_row}:C{tip_row}")
    t2 = wi.cell(row=tip_row, column=1, value="💡  Tips")
    t2.font      = _font(bold=True, size=11, color=C_HEADER_FG)
    t2.fill      = _fill(C_DAY_BG)
    t2.alignment = _align("left")
    wi.row_dimensions[tip_row].height = 22

    tips = [
        "Leave Section and Subject empty to skip a period — the row will be ignored on upload.",
        "Sunday rows are included for extra/special classes only. Leave them blank if unused.",
        "You can delete entire rows for periods you never use to keep the file tidy.",
        "Do NOT change Faculty Username, Day, Period, or Time columns — they are pre-filled.",
        "Upload this file via Department → Timetable → Upload Excel Timetable.",
    ]
    for j, tip in enumerate(tips, start=tip_row + 1):
        wi.merge_cells(f"A{j}:C{j}")
        c = wi.cell(row=j, column=1, value=f"  ✔  {tip}")
        c.font      = _font(size=10, color="1F3864")
        c.fill      = _fill(C_ALT_BG if j % 2 == 0 else C_WHITE)
        c.alignment = _align("left", wrap=True)
        c.border    = _border()
        wi.row_dimensions[j].height = 18

    # ── finalise ──────────────────────────────────────────────────────────────
    # Set active sheet back to data sheet
    wb.active = ws

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f"attachment; filename=timetable_template_{dept}.xlsx"}
    )

@dept_bp.route("/dept/sessions")
def dept_sessions():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    dept = u["department"] if u["role"]=="dept_head" else request.args.get("dept","CSE")
    section = request.args.get("section","")
    date_filter = request.args.get("date","")
    week_offset = int(request.args.get("week", 0))

    # Generate sessions for every faculty in this dept for the requested week
    # so faculty calendar is always in sync with what dept_head sees
    for uname, ud in db["users"].items():
        if ud.get("department") == dept and ud["role"] in ("faculty","dept_head"):
            generate_sessions_for_week(uname, dept, week_offset)

    # Compute week date range for filtering when no specific date_filter provided
    today = datetime.date.today()
    monday = today - datetime.timedelta(days=today.weekday()) + datetime.timedelta(weeks=week_offset)
    sunday = monday + datetime.timedelta(days=6)

    sessions = []
    for s in db["sessions"]:
        if s.get("dept") != dept:
            continue
        if section and s.get("section") != section:
            continue
        if date_filter:
            if s.get("date") != date_filter:
                continue
        else:
            # Filter by week range
            try:
                sess_date = datetime.date.fromisoformat(s["date"])
            except Exception:
                continue
            if not (monday <= sess_date <= sunday):
                continue
        sessions.append(s)

    for s in sessions:
        key = s.get("att_key","")
        recs = db["attendance_records"].get(key, [])
        total = len(db["students"].get(s["section"], []))
        present = sum(1 for r in recs if r["status"] == "Present")
        s["att_count"] = present
        s["total_students"] = total
        s["att_marked"] = len(recs) >= total > 0
        orig_user = db["users"].get(s.get("faculty_username",""), {})
        s["faculty_name"] = orig_user.get("name", s.get("faculty_username",""))
        if s.get("substitute"):
            sub_user = db["users"].get(s["substitute"], {})
            s["substitute_name"] = sub_user.get("name", s["substitute"])
        else:
            s["substitute_name"] = None
    return jsonify({"sessions": sessions})

@dept_bp.route("/dept/assign_substitute", methods=["POST"])
def dept_assign_substitute():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d = request.get_json(silent=True) or {}
    session_id  = d.get("session_id","")
    substitute  = d.get("substitute","")

    sess = next((s for s in db["sessions"] if s["id"] == session_id), None)
    if not sess:
        return jsonify({"error":"Session not found"}),404

    sess["substitute"] = substitute
    save_db()

    # Regenerate sessions for BOTH the original faculty and the substitute
    # so the substitute's calendar immediately reflects the new assignment
    for uname in {sess["faculty_username"], substitute}:
        if not uname:
            continue
        fac_user = db["users"].get(uname, {})
        dept = fac_user.get("department", sess.get("dept", ""))
        for offset in range(-4, 3):
            generate_sessions_for_week(uname, dept, offset)

    return jsonify({"success": True})

@dept_bp.route("/dept/analytics")
def dept_analytics():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    dept    = u["department"] if u["role"]=="dept_head" else request.args.get("dept","CSE")
    section = request.args.get("section","")
    date    = request.args.get("date", datetime.date.today().isoformat())

    total_stu = len(db["students"].get(section,[]))
    summary = []
    for slot in TIME_SLOTS:
        if slot["id"] in BREAK_SLOTS:
            continue
        key = f"{section}::{slot['id']}::{date}"
        recs = db["attendance_records"].get(key, [])
        present = sum(1 for r in recs if r.get("status") == "Present")
        absent  = max(0, total_stu - present)
        summary.append({
            "slot_id":   slot["id"],
            "label":     slot["label"],
            "time":      f"{slot['start']}–{slot['end']}",
            "present":   present,
            "absent":    absent,
            "total":     total_stu,
            "pct":       round(present/total_stu*100,1) if total_stu else 0,
            "records":   recs,
        })
    return jsonify({"summary": summary, "section": section, "date": date, "dept": dept})

@dept_bp.route("/dept/attendance_section")
def dept_attendance_section():
    u = me()
    if not u or u["role"] not in ("admin","dept_head","faculty"):
        return jsonify({"error":"Forbidden"}),403
    section  = request.args.get("section","")
    slot_id  = request.args.get("slot","")
    date_str = request.args.get("date", datetime.date.today().isoformat())
    key = f"{section}::{slot_id}::{date_str}"
    recs = db["attendance_records"].get(key, [])
    all_students = db["students"].get(section, [])
    present_rolls = {r["roll"] for r in recs if r["status"] == "Present"}
    result = []
    for stu in all_students:
        result.append({
            "roll":   stu["roll"],
            "name":   stu["name"],
            "status": "Present" if stu["roll"] in present_rolls else "Absent",
            "photo":  get_student_photo_b64(stu["roll"]),
        })
    return jsonify({"records": result, "section": section, "slot_id": slot_id, "date": date_str})

@dept_bp.route("/dept/update_attendance", methods=["POST"])
def dept_update_attendance():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d        = request.get_json(silent=True) or {}
    section  = d.get("section","")
    slot_id  = d.get("slot_id","")
    date_str = d.get("date", datetime.date.today().isoformat())
    records  = d.get("records",[])

    try:
        att_date = datetime.date.fromisoformat(date_str)
        if att_date > datetime.date.today():
            return jsonify({"error": "Cannot modify attendance for future dates"}), 400

        cutoff   = datetime.datetime.combine(att_date + datetime.timedelta(days=1),
                                              datetime.time(18, 0))
        if datetime.datetime.now() > cutoff:
            return jsonify({"error": "Attendance can only be modified until 6 PM the next day"}), 403
    except Exception:
        pass

    key = f"{section}::{slot_id}::{date_str}"
    # Normalize records to ensure consistent structure
    clean_records = []
    for r in records:
        if not r.get("roll"):
            continue
        clean_records.append({
            "roll":      r["roll"],
            "name":      r.get("name", r["roll"]),
            "status":    r.get("status", "Absent"),
            "time":      datetime.datetime.now().strftime("%H:%M:%S"),
            "marked_by": session.get("username", "dept_head"),
        })
    db["attendance_records"][key] = clean_records
    save_db()
    dept = u.get("department", "")
    rewrite_attendance_csv_for_key(section, dept, slot_id, date_str, clean_records, u.get("name",""))
    return jsonify({"success": True})

@dept_bp.route("/dept/search_attendance")
def dept_search_attendance():
    u = me()
    if not u or u["role"] not in ("admin","dept_head","faculty"):
        return jsonify({"error":"Forbidden"}),403
    query   = request.args.get("q","").strip().lower()
    section = request.args.get("section","")
    dept    = u["department"] if u["role"] in ("dept_head","faculty") else request.args.get("dept","")

    results = {}
    for key, recs in db["attendance_records"].items():
        parts = key.split("::")
        if len(parts) != 3:
            continue
        sec, slot_id, date_str = parts
        if section and sec != section:
            continue
        for r in recs:
            roll = r.get("roll","").lower()
            name = r.get("name","").lower()
            if query and query not in roll and query not in name:
                continue
            results.setdefault(r["roll"], {
                "roll": r["roll"],
                "name": r.get("name",""),
                "photo": get_student_photo_b64(r["roll"]),
                "records": []
            })
            results[r["roll"]]["records"].append({
                "section": sec,
                "slot_id": slot_id,
                "date": date_str,
                "status": r.get("status","Present"),
            })

    return jsonify({"results": list(results.values())})

@dept_bp.route("/dept/upload_students_excel", methods=["POST"])
def dept_upload_students_excel():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error":"openpyxl not installed"}),500

    file    = request.files.get("file")
    section = request.form.get("section","")

    if not file:
        return jsonify({"error":"No file uploaded"}),400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error":"Only .xlsx files accepted"}),400

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        roll_col = next((i for i,h in enumerate(headers) if "roll" in h), None)
        name_col = next((i for i,h in enumerate(headers) if "name" in h), None)
        if roll_col is None:
            return jsonify({"error":"No 'Roll' column found in Excel"}),400

        students = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            roll = str(row[roll_col]).strip() if row[roll_col] else ""
            name = str(row[name_col]).strip() if (name_col is not None and row[name_col]) else roll
            if not roll or roll.lower() == "none":
                continue
            students.append({"roll": roll, "name": name})

        db["students"][section] = students
        save_db()
        return jsonify({"success": True, "count": len(students)})
    except Exception as e:
        return jsonify({"error": str(e)}),500

@dept_bp.route("/dept/upload_excel", methods=["POST"])
def dept_upload_excel():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error":"openpyxl not installed"}),500

    file    = request.files.get("file")
    section = request.form.get("section","")
    slot_id = request.form.get("slot_id","")
    date_str = request.form.get("date", datetime.date.today().isoformat())
    dept    = u["department"] if u["role"]=="dept_head" else request.form.get("dept","")

    if not file:
        return jsonify({"error":"No file uploaded"}),400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error":"Only .xlsx files accepted"}),400

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        roll_col   = next((i for i,h in enumerate(headers) if "roll" in h), None)
        name_col   = next((i for i,h in enumerate(headers) if "name" in h), None)
        status_col = next((i for i,h in enumerate(headers) if "status" in h), None)
        if roll_col is None:
            return jsonify({"error":"No 'Roll' column found"}),400

        stu_map = {s["roll"]: s["name"] for s in db["students"].get(section,[])}
        key = f"{section}::{slot_id}::{date_str}"
        existing_rolls = {r["roll"] for r in db["attendance_records"].get(key,[])}
        db["attendance_records"].setdefault(key, [])
        count = 0
        now = datetime.datetime.now().strftime("%H:%M:%S")
        for row in ws.iter_rows(min_row=2, values_only=True):
            roll = str(row[roll_col]).strip() if row[roll_col] else ""
            if not roll or roll.lower() == "none":
                continue
            name   = str(row[name_col]).strip() if (name_col is not None and row[name_col]) else stu_map.get(roll, roll)
            status = str(row[status_col]).strip() if (status_col is not None and row[status_col]) else "Present"
            if roll not in existing_rolls:
                db["attendance_records"][key].append({"roll":roll,"name":name,"status":status,"time":now,"marked_by":"excel"})
                existing_rolls.add(roll)
                write_attendance_csv(roll, name, section, dept, slot_id, "", u.get("name",""), status)
                count += 1

        save_db()
        return jsonify({"success": True, "marked": count})
    except Exception as e:
        return jsonify({"error": str(e)}),500

@dept_bp.route("/dept/students")
def dept_students():
    u = me()
    if not u or u["role"] not in ("admin","dept_head","faculty"):
        return jsonify({"error":"Forbidden"}),403
    section = request.args.get("section","")
    students = db["students"].get(section,[])
    result = []
    for s in students:
        result.append({**s, "photo": get_student_photo_b64(s["roll"])})
    return jsonify({"students": result})

@dept_bp.route("/dept/save_students", methods=["POST"])
def dept_save_students():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d = request.get_json(silent=True) or {}
    section  = d.get("section","")
    students = d.get("students",[])
    db["students"][section] = students
    save_db()
    return jsonify({"success": True})

@dept_bp.route("/dept/add_student", methods=["POST"])
def dept_add_student():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d = request.get_json(silent=True) or {}
    section = d.get("section","")
    roll    = d.get("roll","").strip()
    name    = d.get("name","").strip()
    if not section or not roll:
        return jsonify({"error":"section and roll required"}),400
    students = db["students"].setdefault(section, [])
    if any(s["roll"] == roll for s in students):
        return jsonify({"error":"Roll number already exists in this section"}),409
    students.append({"roll": roll, "name": name or roll})
    save_db()
    return jsonify({"success": True})

@dept_bp.route("/dept/remove_student", methods=["POST"])
def dept_remove_student():
    u = me()
    if not u or u["role"] not in ("admin","dept_head"):
        return jsonify({"error":"Forbidden"}),403
    d = request.get_json(silent=True) or {}
    section = d.get("section","")
    roll    = d.get("roll","")
    if section in db["students"]:
        db["students"][section] = [s for s in db["students"][section] if s["roll"] != roll]
        save_db()
    return jsonify({"success": True})

@dept_bp.route("/dept/download_attendance")
def dept_download_attendance():
    u = me()
    if not u or u["role"] not in ("admin","dept_head","faculty"):
        return Response("Forbidden",status=403)
    if not OPENPYXL_AVAILABLE:
        return Response("openpyxl not installed",status=500)
    import openpyxl as _openpyxl

    section = request.args.get("section","")
    date    = request.args.get("date", datetime.date.today().isoformat())

    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{section} {date}"
    ws.append(["Roll", "Name", "Period 1", "Period 2", "Period 3",
               "Period 4", "Period 5", "Period 6", "Period 7", "Period 8"])

    all_students = db["students"].get(section, [])
    period_slots = [s for s in TIME_SLOTS if s["id"] not in BREAK_SLOTS]

    for stu in all_students:
        row = [stu["roll"], stu["name"]]
        for slot in period_slots:
            key = f"{section}::{slot['id']}::{date}"
            recs = db["attendance_records"].get(key, [])
            present_rolls = {r["roll"] for r in recs if r["status"] == "Present"}
            row.append("P" if stu["roll"] in present_rolls else "A")
        ws.append(row)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"attendance_{section}_{date}.xlsx"
    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

@dept_bp.route("/dept/download_student_attendance")
def dept_download_student_attendance():
    u = me()
    if not u or u["role"] not in ("admin","dept_head","faculty"):
        return Response("Forbidden",status=403)
    if not OPENPYXL_AVAILABLE:
        return Response("openpyxl not installed",status=500)

    query   = request.args.get("q","").strip().lower()
    section = request.args.get("section","")

    results = {}
    for key, recs in db["attendance_records"].items():
        parts = key.split("::")
        if len(parts) != 3:
            continue
        sec, slot_id, date_str = parts
        if section and sec != section:
            continue
        slot_label = next((s["label"] for s in TIME_SLOTS if s["id"]==slot_id), slot_id)
        for r in recs:
            roll = r.get("roll","").lower()
            name = r.get("name","").lower()
            if query and query not in roll and query not in name:
                continue
            results.setdefault(r["roll"], {
                "roll": r["roll"], "name": r.get("name",""), "records": []
            })
            results[r["roll"]]["records"].append({
                "section": sec, "slot_id": slot_id, "slot_label": slot_label,
                "date": date_str, "status": r.get("status","Present"),
            })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Attendance"
    ws.append(["Roll", "Name", "Section", "Date", "Period", "Status"])
    for stu in results.values():
        for rec in sorted(stu["records"], key=lambda x: (x["date"],x["slot_id"])):
            ws.append([stu["roll"], stu["name"], rec["section"], rec["date"], rec["slot_label"], rec["status"]])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"student_attendance_{query or 'all'}.xlsx"
    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

